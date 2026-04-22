# utils/export_all_to_excel.py
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from PyQt6.QtWidgets import QWidget, QFileDialog, QMessageBox

import utils.session_store as ss

_HDR_FONT   = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
_HDR_FILL   = PatternFill("solid", fgColor="1F497D")
_HDR_ALIGN  = Alignment(horizontal="center", vertical="center", wrap_text=True)
_CELL_ALIGN = Alignment(horizontal="center", vertical="center")
_THIN       = Side(style="thin", color="BFBFBF")
_BORDER     = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


def _auto_width(ws):
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=8)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 35)


def export_lab_to_excel(parent: QWidget, lab_prefix: str, lab_title: str):
    data = ss.get_all_for_prefix(lab_prefix)
    if not data:
        QMessageBox.information(
            parent, "Нет данных",
            "Нет сохранённых данных для экспорта.\n"
            "Откройте нужные подпункты, введите данные и закройте окна."
        )
        return

    path, _ = QFileDialog.getSaveFileName(
        parent, "Сохранить Excel", f"{lab_title}.xlsx", "Excel (*.xlsx)"
    )
    if not path:
        return

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    for key in sorted(data.keys()):
        entry      = data[key]
        sheet_name = entry.get("sheet_name", key)[:31]
        headers    = entry.get("headers", [])
        rows       = entry.get("rows", [])

        ws = wb.create_sheet(title=sheet_name)

        for c, h in enumerate(headers, start=1):
            cell            = ws.cell(row=1, column=c, value=h)
            cell.font       = _HDR_FONT
            cell.fill       = _HDR_FILL
            cell.alignment  = _HDR_ALIGN
            cell.border     = _BORDER
        ws.row_dimensions[1].height = 38

        for r, row_data in enumerate(rows, start=2):
            for c, val in enumerate(row_data, start=1):
                cell           = ws.cell(row=r, column=c)
                cell.alignment = _CELL_ALIGN
                cell.border    = _BORDER
                if val:
                    try:
                        cell.value         = float(val.replace(",", "."))
                        cell.number_format = "0.####"
                    except (ValueError, AttributeError):
                        cell.value = val
            ws.row_dimensions[r].height = 18

        _auto_width(ws)
        ws.freeze_panes = "A2"

    # Лист-оглавление
    idx = wb.create_sheet(title="Оглавление", index=0)
    idx["B1"] = lab_title
    idx["B1"].font = Font(name="Calibri", bold=True, size=13)
    for i, key in enumerate(sorted(data.keys()), start=3):
        sn   = data[key].get("sheet_name", key)[:31]
        cell = idx.cell(row=i, column=2, value=sn)
        cell.hyperlink = f"'{sn}'!A1"
        cell.font      = Font(name="Calibri", color="0563C1", underline="single")
    idx.column_dimensions["B"].width = 45

    wb.save(path)
    QMessageBox.information(
        parent, "Экспорт завершён",
        f"Файл сохранён:\n{path}\n\nЛистов: {len(data)}"
    )